from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import transaction
from .forms import NetToGrossForm
from .utils import calculate_basic_from_net, calculate_primes_automatiques, calculate_exempt_primes_amounts
from .models import Employee
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

@login_required
def net_to_gross_view(request):
    result = None
    primes_auto = None
    employees = Employee.objects.all()[:10]  # Derniers 10 employés pour affichage
    
    if request.method == "POST":
        form = NetToGrossForm(request.POST)
        if form.is_valid():
            nom_complet = form.cleaned_data['nom_complet']
            net_salary = form.cleaned_data['net_salary']
            has_exempt_primes = form.cleaned_data.get('has_exempt_primes', False)
            prime_retraite = form.cleaned_data.get('prime_retraite', False)
            prime_interim = form.cleaned_data.get('prime_interim', False)
            prime_anciennete = form.cleaned_data.get('prime_anciennete', False)
            
            # Calculer automatiquement les primes taxables
            primes_auto = calculate_primes_automatiques(net_salary)
            
            # Calculer le total des primes taxables
            primes_taxables = (
                primes_auto['prime_cherte_vie'] +
                primes_auto['prime_craie'] +
                primes_auto['indemnite_logement'] +
                primes_auto['indemnite_transport'] +
                primes_auto['indemnite_repas'] +
                primes_auto['autre_gratification']
            )
            
            # Calculer les primes exonérées si sélectionnées
            primes_exonerees = 0
            exempt_primes_amounts = {}
            if has_exempt_primes:
                selected_primes = []
                if prime_retraite:
                    selected_primes.append('retraite')
                if prime_interim:
                    selected_primes.append('interim')
                if prime_anciennete:
                    selected_primes.append('anciennete')
                
                exempt_primes_amounts = calculate_exempt_primes_amounts(net_salary, selected_primes)
                primes_exonerees = sum(exempt_primes_amounts.values())
            
            # Calculer sans avantages ni déductions (0 pour les deux)
            result = calculate_basic_from_net(
                net_salary, 
                0,  # Pas d'avantages généraux
                0,  # Pas de déductions
                primes_taxables,
                primes_exonerees
            )
            
            # Sauvegarder automatiquement l'employé
            try:
                with transaction.atomic():
                    employee = Employee.objects.create(
                        nom_complet=nom_complet,
                        salaire_net=net_salary,
                        salaire_base=result['basic'],
                        salaire_brut=result['gross'],
                        salaire_imposable=result['imposable'],
                        cnss_employe=result['cnss'],
                        rts=result['rts'],
                        total_charges_employee=result['total_charges_employee'],
                        cnss_employeur=result['cnss_employer'],
                        versement_forfaitaire=result['versement_forfaitaire'],
                        taxe_apprentissage=result['taxe_apprentissage'],
                        total_cnss_patronal=result['total_cnss_patronal'],
                        # Primes taxables détaillées
                        prime_cherte_vie=primes_auto['prime_cherte_vie'],
                        prime_craie=primes_auto['prime_craie'],
                        indemnite_logement=primes_auto['indemnite_logement'],
                        indemnite_transport=primes_auto['indemnite_transport'],
                        indemnite_repas=primes_auto['indemnite_repas'],
                        autre_gratification=primes_auto['autre_gratification'],
                        primes_taxables=primes_taxables,
                        # Primes exonérées
                        prime_retraite=exempt_primes_amounts.get('prime_retraite', 0),
                        prime_interim=exempt_primes_amounts.get('prime_interim', 0),
                        prime_anciennete=exempt_primes_amounts.get('prime_anciennete', 0),
                        primes_exonerees=primes_exonerees,
                        ecart_imposable=result['ecart_imposable']
                    )
                    
                    messages.success(request, f"✅ Employé '{nom_complet}' ajouté avec succès !")
                    # Recharger la liste des employés
                    employees = Employee.objects.all()[:10]
                    
            except Exception as e:
                messages.error(request, f"❌ Erreur lors de l'ajout : {str(e)}")
    else:
        form = NetToGrossForm()

    return render(request, "salary/index.html", {
        "form": form, 
        "result": result,
        "primes_auto": primes_auto,
        "exempt_primes_amounts": exempt_primes_amounts if 'exempt_primes_amounts' in locals() else {},
        "employees": employees
    })

@login_required
def export_excel_view(request):
    """Exporter la liste des employés en Excel"""
    employees = Employee.objects.all()
    
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des Employés"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # En-têtes selon l'ordre demandé
    headers = [
        # 1. Nom
        "Nom Complet",
        # 2. Salaire de base
        "Salaire Base",
        # 3. Toutes les primes (exonérées et non exonérées)
        "Prime Cherté de Vie", "Prime Craie", "Indemnité Logement", 
        "Indemnité Transport", "Indemnité Repas", "Autre Gratification",
        "Prime Retraite", "Prime Intérim", "Prime Ancienneté",
        # 4. Salaire brut
        "Salaire Brut",
        # 5. CNSS employé et employeur
        "CNSS Employé", "CNSS Employeur",
        # 6. Écart imposable
        "Écart Imposable",
        # 7. Salaire imposable
        "Salaire Imposable",
        # 8. Détails RTS
        "RTS Tranche 1 (0%)", "RTS Tranche 2 (5%)", "RTS Tranche 3 (8%)", 
        "RTS Tranche 4 (10%)", "RTS Tranche 5 (15%)", "RTS Tranche 6 (20%)",
        # 9. RTS total
        "RTS Total",
        # 10. Total cotisations employé et employeur
        "Total Charges Employé", "Total CNSS Patronal",
        # 11. Versement forfaitaire et taxe d'apprentissage
        "Versement Forfaitaire", "Taxe Apprentissage",
        # 12. Salaire net (dernière colonne)
        "Salaire Net"
    ]
    
    # Écrire les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Écrire les données selon le nouvel ordre
    for row, employee in enumerate(employees, 2):
        col = 1
        
        # 1. Nom Complet
        ws.cell(row=row, column=col, value=employee.nom_complet)
        col += 1
        
        # 2. Salaire Base
        ws.cell(row=row, column=col, value=float(employee.salaire_base))
        col += 1
        
        # 3. Toutes les primes (non exonérées d'abord)
        ws.cell(row=row, column=col, value=float(employee.prime_cherte_vie))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.prime_craie))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.indemnite_logement))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.indemnite_transport))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.indemnite_repas))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.autre_gratification))
        col += 1
        
        # Primes exonérées
        ws.cell(row=row, column=col, value=float(employee.prime_retraite))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.prime_interim))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.prime_anciennete))
        col += 1
        
        # 4. Salaire Brut
        ws.cell(row=row, column=col, value=float(employee.salaire_brut))
        col += 1
        
        # 5. CNSS Employé et Employeur
        ws.cell(row=row, column=col, value=float(employee.cnss_employe))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.cnss_employeur))
        col += 1
        
        # 6. Écart Imposable
        ws.cell(row=row, column=col, value=float(employee.ecart_imposable))
        col += 1
        
        # 7. Salaire Imposable
        ws.cell(row=row, column=col, value=float(employee.salaire_imposable))
        col += 1
        
        # 8. Détails RTS par tranche
        from .utils import calculate_rts_detailed
        _, rts_details = calculate_rts_detailed(employee.salaire_imposable)
        
        # Initialiser toutes les tranches à 0
        tranche_amounts = [0, 0, 0, 0, 0, 0]  # 6 tranches
        
        # Extraire les montants des détails
        import re
        for detail in rts_details:
            if "=" in detail and "TOTAL" not in detail:
                try:
                    # Extraire le montant de la chaîne
                    amount_str = detail.split("=")[-1].strip().replace(" GNF", "").replace(" ", "").replace(",", "")
                    amount = float(amount_str) if amount_str else 0
                    
                    # Déterminer quelle tranche basée sur le pourcentage avec regex
                    if re.search(r'^0%', detail):
                        tranche_amounts[0] = amount
                    elif re.search(r'^5%', detail):
                        tranche_amounts[1] = amount
                    elif re.search(r'^8%', detail):
                        tranche_amounts[2] = amount
                    elif re.search(r'^10%', detail):
                        tranche_amounts[3] = amount
                    elif re.search(r'^15%', detail):
                        tranche_amounts[4] = amount
                    elif re.search(r'^20%', detail):
                        tranche_amounts[5] = amount
                except (ValueError, IndexError):
                    continue
        
        # Écrire les 6 tranches RTS
        for i, amount in enumerate(tranche_amounts):
            ws.cell(row=row, column=col, value=amount)
            col += 1
        
        # 9. RTS Total
        ws.cell(row=row, column=col, value=float(employee.rts))
        col += 1
        
        # 10. Total cotisations employé et employeur
        ws.cell(row=row, column=col, value=float(employee.total_charges_employee))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.total_cnss_patronal))
        col += 1
        
        # 11. Versement forfaitaire et taxe d'apprentissage
        ws.cell(row=row, column=col, value=float(employee.versement_forfaitaire))
        col += 1
        ws.cell(row=row, column=col, value=float(employee.taxe_apprentissage))
        col += 1
        
        # 12. Salaire Net (dernière colonne)
        ws.cell(row=row, column=col, value=float(employee.salaire_net))
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20
    
    # Sauvegarder dans un buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Créer la réponse HTTP
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="liste_employes.xlsx"'
    
    return response

@login_required
def delete_all_employees_view(request):
    """Supprimer tous les employés"""
    if request.method == "POST":
        try:
            count = Employee.objects.count()
            Employee.objects.all().delete()
            messages.success(request, f"✅ {count} employé(s) supprimé(s) avec succès !")
        except Exception as e:
            messages.error(request, f"❌ Erreur lors de la suppression : {str(e)}")
    
    return redirect('index')
